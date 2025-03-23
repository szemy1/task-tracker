# exports/excel_exporter.py

import os
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime

EXPORTS_FOLDER = "exports"

def export_tasks_to_excel(tasks):
    os.makedirs(EXPORTS_FOLDER, exist_ok=True)
    now_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"tasks_export_{now_str}.xlsx"
    filepath = os.path.join(EXPORTS_FOLDER, filename)

    wb = Workbook()

    # === 1. Munkalap: Feladatok összesítve ===
    ws1 = wb.active
    ws1.title = "Feladatok"

    headers = ["Cím", "Leírás", "Kezdés", "Befejezés", "Időtartam", "Záró megjegyzés"]
    ws1.append(headers)

    for cell in ws1[1]:
        cell.font = Font(bold=True)

    for task in tasks:
        duration = task.get_duration()
        duration_str = str(duration) if duration else "–"

        close_note = ""
        for ts, msg in reversed(task.logs):
            if msg.startswith("Záró megjegyzés:"):
                close_note = msg.replace("Záró megjegyzés: ", "")
                break

        ws1.append([
            task.title,
            task.description,
            str(task.start_time) if task.start_time else "–",
            str(task.end_time) if task.end_time else "–",
            duration_str,
            close_note
        ])

    ws1.auto_filter.ref = f"A1:F{ws1.max_row}"
    for col in ws1.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws1.column_dimensions[col[0].column_letter].width = max_length + 2

    # === 2. Munkalap: Részletes logok ===
    ws2 = wb.create_sheet(title="Részletes napló")

    ws2.append(["Feladat", "Időbélyeg", "Esemény"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)

    for task in tasks:
        for ts, msg in task.logs:
            ws2.append([
                task.title,
                str(ts),
                msg
            ])

    wb.save(filepath)
    return filepath
